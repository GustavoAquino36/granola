"""
Importador MV Krupp.xlsx → Granola CRM
Importa: Processos, Arquivo, Leads, Financeiro, Fundos
Ignora: Logins e Senhas, Planilhas, Parcerias, Casos sem Sequência
"""
import sys, os, re
from pathlib import Path
from datetime import datetime

# Fix path
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from granola.database import GranolaDB, AuthDB, init_db

XLSX = Path(os.environ.get("XLSX_PATH", r"C:\Users\zoval\Downloads\MV Krupp.xlsx"))

# ============================================================
#  Helpers
# ============================================================
_clientes_cache = {}  # nome_lower -> id

def get_or_create_cliente(db, nome, telefone=None, observacao=None):
    """Retorna ID do cliente, criando se necessário."""
    if not nome or not nome.strip():
        return None
    nome = nome.strip()
    key = nome.lower()
    if key in _clientes_cache:
        return _clientes_cache[key]
    # Busca existente
    rows = db.listar_clientes(busca=nome, limite=5)
    for r in rows:
        if r["nome"].lower() == key:
            _clientes_cache[key] = r["id"]
            return r["id"]
    # Cria novo
    data = {"nome": nome, "tipo": "PF"}
    if telefone:
        data["telefone"] = str(telefone).strip()
    if observacao:
        data["observacao"] = observacao
    cid = db.upsert_cliente(data)
    _clientes_cache[key] = cid
    print(f"  + Cliente: {nome} (id={cid})")
    return cid


def parse_valor(v):
    """Converte valor da planilha (str ou float) para float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # Remove R$, espaços, pontos de milhar
    s = s.replace("R$", "").replace(" ", "").strip()
    if not s or s == "-":
        return 0.0
    # Trata formato brasileiro: 35.475.00 ou 354.750,00
    # Se tem vírgula, é separador decimal brasileiro
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # Pode ser 35475.00 (decimal com ponto) — manter
        pass
    try:
        return float(s)
    except ValueError:
        return 0.0


def normalize_area(area_raw):
    """Normaliza área do processo."""
    if not area_raw:
        return "civel"
    a = area_raw.strip().lower()
    if "trabalhist" in a:
        return "trabalhista"
    if "criminal" in a or "penal" in a:
        return "criminal"
    if "previd" in a or "inss" in a:
        return "previdenciario"
    if "tribut" in a:
        return "tributario"
    if "rpv" in a:
        return "civel"
    return "civel"


def normalize_status(obs, mov):
    """Tenta inferir status do processo pelas observações."""
    text = f"{obs or ''} {mov or ''}".lower()
    if "extint" in text or "quitad" in text:
        return "encerrado"
    if "arquivad" in text:
        return "arquivado"
    if "suspens" in text:
        return "suspenso"
    return "ativo"


# ============================================================
#  Importadores por aba
# ============================================================

def importar_processos(db, ws, arquivado=False):
    """Importa aba Processos ou Arquivo."""
    sheet_name = "Arquivo" if arquivado else "Processos"
    count = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        area_raw = row[0]
        numero = row[1]
        # link = row[2]  # não usado
        autor = row[3]
        reu = row[4]
        tipo = row[5]
        honorarios = row[6]
        valor_causa_raw = row[7]
        sucumbencia = row[8]
        mov = row[9]
        teor = row[10]
        obs = row[11]

        # Skip linhas vazias
        if not numero and not autor:
            continue

        # Limpa
        numero = str(numero).strip() if numero else None
        autor = str(autor).strip() if autor else None
        reu = str(reu).strip() if reu else None

        if not numero and not autor:
            continue

        # Monta observação completa
        obs_parts = []
        if honorarios:
            obs_parts.append(f"Honorários: {honorarios}")
        if sucumbencia:
            obs_parts.append(f"Sucumbência: {sucumbencia}")
        if teor:
            obs_parts.append(f"Teor: {teor}")
        if obs:
            obs_parts.append(str(obs))
        observacao = " | ".join(obs_parts) if obs_parts else None

        # Status
        if arquivado:
            status = "arquivado"
        else:
            status = normalize_status(obs, mov)

        # Valor causa
        valor_causa = parse_valor(valor_causa_raw)

        # Título
        titulo = str(tipo).strip() if tipo else (f"Processo {numero}" if numero else "Sem título")

        # Área
        area = normalize_area(area_raw)

        # Verifica se processo já existe (por CNJ)
        if numero:
            existing = db.listar_processos(busca=numero, limite=1)
            if existing:
                print(f"  ~ Processo já existe: {numero} — skip")
                continue

        # Detecta qual das partes é cliente do escritório.
        # Heurística: em trabalhista, empresas (razão social com LTDA/S.A./EIRELI/ME)
        # tendem a ser reclamadas; mas aqui MV Krupp pode representar ambos os lados.
        # Regra adotada: se "autor" tem marcadores de PJ → provavelmente é a empresa
        # cliente (MV Krupp representa empresas). Caso contrário, usa autor.
        def _eh_pj(nome):
            if not nome:
                return False
            up = nome.upper()
            return any(tag in up for tag in (" LTDA", " S.A", " S/A", " EIRELI", " ME ", " EPP", "SOCIEDADE", "ASSOCIAC"))

        cliente_nome = autor
        parte_nome = reu
        polo_cliente = "ativo"
        # Em trabalhista, se o réu é PJ e o autor é PF, provavelmente a empresa é cliente
        if area == "trabalhista" and reu and _eh_pj(reu) and autor and not _eh_pj(autor):
            cliente_nome = reu
            parte_nome = autor
            polo_cliente = "passivo"

        cliente_id = get_or_create_cliente(db, cliente_nome)

        # Cria processo
        proc_data = {
            "cliente_id": cliente_id,
            "numero_cnj": numero,
            "titulo": titulo,
            "area": area,
            "status": status,
            "valor_causa": valor_causa,
            "parte_contraria": parte_nome,
            "polo": polo_cliente,
            "observacao": observacao,
        }
        pid = db.upsert_processo(proc_data)
        count += 1
        print(f"  + [{sheet_name}] {numero or 'sem-cnj'} — {titulo} (id={pid}) [{polo_cliente}]")

        # Cria movimentação se tiver
        if mov and str(mov).strip() and str(mov).strip() != "-":
            mov_text = str(mov).strip()
            # Tenta extrair data do início da movimentação
            data_mov = datetime.now().strftime("%Y-%m-%d")
            date_match = re.match(r"(\d{2}/\d{2}/\d{4})", mov_text)
            if date_match:
                try:
                    dt = datetime.strptime(date_match.group(1), "%d/%m/%Y")
                    data_mov = dt.strftime("%Y-%m-%d")
                except ValueError:
                    pass
            try:
                db.criar_movimentacao({
                    "processo_id": pid,
                    "descricao": mov_text,
                    "data_movimento": data_mov,
                    "tipo": "despacho",
                    "fonte": "importacao",
                })
            except Exception:
                pass  # Duplicata ou erro — ignora

        # Cria ambas as partes (autor e réu) em granola_partes
        if autor:
            try:
                db.upsert_parte({
                    "processo_id": pid,
                    "nome": autor,
                    "tipo": "autor",
                    "polo": "ativo",
                })
            except Exception:
                pass
        if reu:
            try:
                db.upsert_parte({
                    "processo_id": pid,
                    "nome": reu,
                    "tipo": "reu",
                    "polo": "passivo",
                })
            except Exception:
                pass

    print(f"  >> {sheet_name}: {count} processos importados")
    return count


def importar_leads(db, ws):
    """Importa aba Leads como clientes com observações."""
    count = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        cliente = row[0]
        area = row[1]
        responsavel = row[2]
        telefone = row[3]
        mov = row[4]
        pontuacao = row[5]
        bola_vez = row[6]

        if not cliente or not str(cliente).strip():
            continue

        nome = str(cliente).strip()

        # Monta observação
        obs_parts = [f"[LEAD]"]
        if area:
            obs_parts.append(f"Área: {area}")
        if responsavel:
            obs_parts.append(f"Responsável: {responsavel}")
        if mov:
            obs_parts.append(f"Último contato: {mov}")
        if pontuacao:
            obs_parts.append(f"Status: {pontuacao}")
        if bola_vez:
            obs_parts.append(f"Próximo passo: {bola_vez}")
        observacao = " | ".join(obs_parts)

        tel = str(telefone).strip() if telefone and str(telefone).strip() not in ("-", "?") else None

        cid = get_or_create_cliente(db, nome, telefone=tel, observacao=observacao)
        count += 1

    print(f"  >> Leads: {count} clientes importados/atualizados")
    return count


def importar_financeiro(db, ws):
    """Importa aba Financeiro — despesas fixas e futuras."""
    count = 0

    # Seção 1: Valores Fixos (rows 4-16, col A=origem, C=vencimento dia, D=valor)
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=16, values_only=True), 4):
        origem = row[0]
        contrato = row[1]
        venc_dia = row[2]
        valor_raw = row[3]

        if not origem or not str(origem).strip() or str(origem).strip() in ("TOTAIS", "Variável", "Resultado final"):
            continue

        valor = parse_valor(valor_raw)
        descricao = f"Despesa fixa — {str(origem).strip()}"
        if contrato:
            descricao += f" (contrato: {contrato})"

        # Data vencimento: dia do mês atual
        dia = int(venc_dia) if venc_dia and isinstance(venc_dia, (int, float)) else 1
        hoje = datetime.now()
        try:
            data_venc = hoje.replace(day=dia).strftime("%Y-%m-%d")
        except ValueError:
            data_venc = hoje.strftime("%Y-%m-%d")

        fid = db.upsert_financeiro({
            "tipo": "despesa",
            "categoria": "fixo",
            "descricao": descricao,
            "valor": valor if valor > 0 else 0,
            "data_vencimento": data_venc,
            "status": "pendente",
        })
        count += 1
        print(f"  + [Fin] {descricao} = R${valor:.2f}")

    # Seção 2: Valores Futuros (rows 4-10, col F=origem, G=parcelas, H=venc, I ou J=valor)
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=10, values_only=True), 4):
        origem_fut = row[5]
        parcelas = row[6]
        venc_fut = row[7]
        valor1 = row[8]
        valor2 = row[9]

        if not origem_fut or not str(origem_fut).strip():
            continue

        valor = parse_valor(valor1) or parse_valor(valor2)
        if valor <= 0:
            continue

        descricao = f"Futuro — {str(origem_fut).strip()}"
        if parcelas:
            descricao += f" ({int(parcelas)}x)"

        # Parse data DD.MM.YY
        data_venc = datetime.now().strftime("%Y-%m-%d")
        if venc_fut:
            try:
                dt = datetime.strptime(str(venc_fut).strip(), "%d.%m.%y")
                data_venc = dt.strftime("%Y-%m-%d")
            except ValueError:
                pass

        fid = db.upsert_financeiro({
            "tipo": "receita",
            "categoria": "futuro",
            "descricao": descricao,
            "valor": valor,
            "data_vencimento": data_venc,
            "status": "pendente",
        })
        count += 1
        print(f"  + [Fin] {descricao} = R${valor:.2f}")

    # Seção 3: Despesas fixas do escritório (rows 23-30+, col A=origem, C=venc dia)
    for i, row in enumerate(ws.iter_rows(min_row=24, max_row=40, values_only=True), 24):
        origem = row[0]
        pagante = row[1]
        venc_dia = row[2]
        valor_raw = row[3]

        if not origem or not str(origem).strip():
            continue

        valor = parse_valor(valor_raw)
        descricao = f"Custo fixo — {str(origem).strip()}"
        if pagante:
            descricao += f" ({pagante})"

        dia = int(venc_dia) if venc_dia and isinstance(venc_dia, (int, float)) else 1
        hoje = datetime.now()
        try:
            data_venc = hoje.replace(day=min(dia, 28)).strftime("%Y-%m-%d")
        except ValueError:
            data_venc = hoje.strftime("%Y-%m-%d")

        fid = db.upsert_financeiro({
            "tipo": "despesa",
            "categoria": "custo_escritorio",
            "descricao": descricao,
            "valor": valor if valor > 0 else 0,
            "data_vencimento": data_venc,
            "status": "pendente",
        })
        count += 1
        print(f"  + [Fin] {descricao}")

    print(f"  >> Financeiro: {count} lançamentos importados")
    return count


def importar_fundos(db, ws):
    """Importa aba Fundos — processos de precatórios."""
    count = 0
    # Parte 1: Processos dos clientes (rows 2-13)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=14, values_only=True), 2):
        processo = row[0]
        fase = row[1]
        autor = row[2]
        reu = row[3]
        comprador = row[4]
        cnpj = row[5]

        if not processo or not str(processo).strip():
            continue

        numero = str(processo).strip()
        autor_nome = str(autor).strip() if autor else None
        reu_nome = str(reu).strip() if reu else None

        # Cria cliente
        cliente_id = get_or_create_cliente(db, autor_nome) if autor_nome else None

        # Mapeia fase
        area = "previdenciario" if reu_nome and "INSS" in reu_nome else "civel"
        titulo = f"Fundo — {fase or 'N/A'}"

        # Verifica duplicata
        existing = db.listar_processos(busca=numero, limite=1)
        if existing:
            continue

        obs_parts = []
        if comprador and str(comprador).strip() not in ("-", ""):
            obs_parts.append(f"Comprador: {comprador}")
        if cnpj and str(cnpj).strip() not in ("-", ""):
            obs_parts.append(f"CNPJ: {cnpj}")

        pid = db.upsert_processo({
            "cliente_id": cliente_id,
            "numero_cnj": numero,
            "titulo": titulo,
            "area": area,
            "status": "ativo",
            "parte_contraria": reu_nome,
            "polo": "ativo",
            "observacao": " | ".join(obs_parts) if obs_parts else None,
        })
        count += 1
        print(f"  + [Fundos] {numero} — {autor_nome} vs {reu_nome}")

    print(f"  >> Fundos: {count} processos importados")
    return count


# ============================================================
#  Main
# ============================================================
def main():
    print("=" * 50)
    print("  IMPORTADOR MV Krupp -> Granola CRM")
    print("=" * 50)
    print()

    if not XLSX.exists():
        print(f"ERRO: Arquivo não encontrado: {XLSX}")
        sys.exit(1)

    # Init DB
    init_db()
    db = GranolaDB()

    # Abrir Excel
    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    print(f"Abas encontradas: {wb.sheetnames}")
    print()

    total = 0

    # 1. Processos ativos
    if "Processos" in wb.sheetnames:
        print("[1/5] Importando Processos...")
        total += importar_processos(db, wb["Processos"], arquivado=False)
        print()

    # 2. Arquivo (processos encerrados)
    if "Arquivo" in wb.sheetnames:
        print("[2/5] Importando Arquivo...")
        total += importar_processos(db, wb["Arquivo"], arquivado=True)
        print()

    # 3. Leads
    if "Leads" in wb.sheetnames:
        print("[3/5] Importando Leads...")
        total += importar_leads(db, wb["Leads"])
        print()

    # 4. Financeiro
    if "Financeiro" in wb.sheetnames:
        print("[4/5] Importando Financeiro...")
        total += importar_financeiro(db, wb["Financeiro"])
        print()

    # 5. Fundos
    if "Fundos" in wb.sheetnames:
        print("[5/5] Importando Fundos...")
        total += importar_fundos(db, wb["Fundos"])
        print()

    db.close()

    print("=" * 50)
    print(f"  IMPORTAÇÃO CONCLUÍDA — {total} registros")
    print(f"  Banco: granola/data/granola.db")
    print("=" * 50)


if __name__ == "__main__":
    main()
